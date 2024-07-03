import argparse
import os
import uuid
from pathlib import Path
from typing import Dict, List

from lxml import etree
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from reqif.experimental.reqif_schema import ReqIFSchema
from reqif.helpers.string.xhtml_indent import reqif_indent_xhtml_string
from reqif.models.reqif_core_content import ReqIFCoreContent
from reqif.models.reqif_data_type import ReqIFDataTypeDefinitionString, \
    ReqIFDataTypeDefinitionXHTML, ReqIFDataTypeDefinitionDateIdentifier, \
    ReqIFDataTypeDefinitionEnumeration
from reqif.models.reqif_namespace_info import ReqIFNamespaceInfo
from reqif.models.reqif_req_if_content import ReqIFReqIFContent
from reqif.models.reqif_reqif_header import ReqIFReqIFHeader
from reqif.models.reqif_spec_hierarchy import ReqIFSpecHierarchy
from reqif.models.reqif_spec_object import ReqIFSpecObject, SpecObjectAttribute
from reqif.models.reqif_spec_object_type import (
    ReqIFSpecObjectType,
    SpecAttributeDefinition,
)
from reqif.models.reqif_specification import ReqIFSpecification
from reqif.models.reqif_specification_type import ReqIFSpecificationType
from reqif.models.reqif_types import SpecObjectAttributeType
from reqif.object_lookup import ReqIFObjectLookup
from reqif.parser import ReqIFParser
from reqif.reqif_bundle import ReqIFBundle
from reqif.unparser import ReqIFUnparser


# "_" is needed to make the ReqIF XML schema validator happy.
def create_uuid():
    return "_" + uuid.uuid4().hex


def _convert_spec_object_to_node(
    spec_object: ReqIFSpecObject,
    reqif_bundle: ReqIFBundle,
    reqif_schema: ReqIFSchema,
    level,
):
    assert 1 <= level <= 10, "Expecting a reasonable level of nesting."
    spec_object_type_ref = spec_object.spec_object_type

    spec_object_type = reqif_bundle.get_spec_object_type_by_ref(
        spec_object_type_ref
    )
    assert spec_object_type is not None
    row_dict = {}
    for spec_object_attribute_ in spec_object.attributes:
        attribute_definition = reqif_schema.spec_object_type_attributes[
            spec_object_attribute_.definition_ref
        ]
        assert attribute_definition.long_name is not None

        data_type = reqif_schema.data_type_definitions[
            attribute_definition.datatype_definition
        ]
        assert data_type is not None
        if isinstance(data_type, ReqIFDataTypeDefinitionDateIdentifier):
            assert isinstance(spec_object_attribute_.value, str)
            row_dict[
                attribute_definition.long_name
            ] = spec_object_attribute_.value
        elif isinstance(data_type, ReqIFDataTypeDefinitionEnumeration):
            assert isinstance(spec_object_attribute_.value, list)

            enum_values: List[str] = []
            for enum_value_identifier_ in spec_object_attribute_.value:
                assert data_type.values is not None
                for data_type_enum_value_ in data_type.values:
                    if (
                        data_type_enum_value_.identifier
                        == enum_value_identifier_
                    ):
                        assert data_type_enum_value_.long_name is not None
                        enum_values.append(data_type_enum_value_.long_name)
                        break
                else:
                    raise AssertionError("Enum value not found.")
            row_dict[attribute_definition.long_name] = ", ".join(
                enum_values
            )
        elif isinstance(data_type, ReqIFDataTypeDefinitionString):
            assert isinstance(spec_object_attribute_.value, str)
            row_dict[
                attribute_definition.long_name
            ] = spec_object_attribute_.value
        elif isinstance(data_type, ReqIFDataTypeDefinitionXHTML):
            assert isinstance(spec_object_attribute_.value, str)
            row_dict[
                attribute_definition.long_name
            ] = spec_object_attribute_.value
        else:
            raise NotImplementedError(data_type)

    return row_dict


def convert_excel_to_reqif(path_to_excel: str) -> str:
    """
    This conversion script only works against an Excel file supplied by a user.
    This Excel file features a very basic field schema that has the following
    traits:
    - A node/object type column is not available, so this script treats all nodes
      as Requirements. An outcome of this simplicity is that the exported document
      has no Sections/Chapters, and the list of requirements is plain.
    - Relations are stored in the first column called Inlinks. This script does
      not convert these relations to ReqIF Spec Relations yet.
      FIXME: Fix this.
    """
    workbook = load_workbook(filename=path_to_excel)

    sheet = workbook.active

    date_now = "2023-01-01T00:00:00.000+02:00"

    """
    Create a single String data type. It will be used for all
    Spec Attribute Definitions below.
    """
    string_data_type = ReqIFDataTypeDefinitionString(
        identifier=create_uuid(),
        last_change=date_now,
        long_name="String type",
        max_length="50",
    )

    """
    Use the first row's column titles to create Spec Attribute Definitions which
    are essentially requirement fields.
    """
    requirement_spec_attributes: List[SpecAttributeDefinition] = []
    map_column_titles_to_requirement_spec_attributes: Dict[str, SpecAttributeDefinition] = {}

    # FIXME: This Excel has 13 requirement fields.
    # Some fields further right contain irregular notes and meta information
    # about the document.
    total_requirement_fields = 13
    column_titles: List[str] = next(sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=total_requirement_fields, values_only=True))
    for column_title_ in column_titles:
        requirement_text_attribute = SpecAttributeDefinition(
            attribute_type=SpecObjectAttributeType.XHTML,
            identifier=create_uuid(),
            datatype_definition=string_data_type.identifier,
            last_change=date_now,
            long_name=column_title_,
        )
        requirement_spec_attributes.append(requirement_text_attribute)
        map_column_titles_to_requirement_spec_attributes[
            column_title_
        ] = requirement_text_attribute

    """
    Create a Spec Object Type for a single Requirement type.
    """
    spec_object_type = ReqIFSpecObjectType.create(
        identifier=create_uuid(),
        long_name="Requirement",
        last_change=date_now,
        attribute_definitions=requirement_spec_attributes,
    )

    """
    Iterate over all rows of the Excel file and create at the same time:
    - Spec Objects which are single requirements that contain the actual row values.
    - Spec Hierarchies which give the Spec Objects their place in the document.
    """
    spec_objects: List[ReqIFSpecObject] = []
    spec_hierarchies: List[ReqIFSpecHierarchy] = []
    for row_ in sheet.iter_rows(
        min_row=2, min_col=1, max_col=10, values_only=True
    ):
        spec_object_attributes: List[SpecObjectAttribute] = []

        for row_value_idx_, row_value_ in enumerate(row_):
            # Empty cells are ignored.
            if row_value_ is None:
                continue

            assert row_value_ is not None
            column_field = column_titles[row_value_idx_]
            spec_object_attribute = map_column_titles_to_requirement_spec_attributes[
                column_field
            ]
            row_value_ = f"<xhtml:div>{row_value_.strip()}</xhtml:div>"
            spec_object_attribute = SpecObjectAttribute(
                attribute_type=SpecObjectAttributeType.XHTML,
                definition_ref=spec_object_attribute.identifier,
                value=reqif_indent_xhtml_string(row_value_),
            )
            spec_object_attributes.append(spec_object_attribute)

        spec_object = ReqIFSpecObject(
            identifier=create_uuid(),
            attributes=spec_object_attributes,
            spec_object_type=spec_object_type.identifier,
            last_change=date_now,
        )
        spec_objects.append(spec_object)

        spec_hierarchy = ReqIFSpecHierarchy(
            xml_node=None,
            is_self_closed=False,
            identifier=create_uuid(),
            last_change=date_now,
            long_name=None,
            spec_object=spec_object.identifier,
            children=[],
            level=1,
        )
        spec_hierarchies.append(spec_hierarchy)

    """
    Create a single specification type which is always needed even if there is
    only one document.
    """
    specification_type = ReqIFSpecificationType(
        identifier=create_uuid(),
        description=None,
        last_change=date_now,
        long_name="Software Requirements Specification Document",
        spec_attributes=None,
        spec_attribute_map={},
        is_self_closed=True,
    )

    """
    Create a single specification which is of the Specification Type created just before.
    Specification corresponds to a document.
    """
    specification = ReqIFSpecification(
        identifier=create_uuid(),
        last_change=date_now,
        long_name="ReqIF library requirements specification",
        # Empty <VALUES/> tag is needed to make the XML schema validator happy.
        values=[],
        specification_type=specification_type.identifier,
        children=spec_hierarchies,
    )

    """
    Create the ReqIF's ReqIF Content class that stores together all types,
    objects, and the specification.
    """
    reqif_content = ReqIFReqIFContent(
        data_types=[string_data_type],
        spec_types=[
            specification_type,
            spec_object_type,
        ],
        spec_objects=spec_objects,
        spec_relations=[],
        specifications=[specification],
        spec_relation_groups=[],
    )

    core_content = ReqIFCoreContent(req_if_content=reqif_content)

    """
    Create the ReqIF's Namespace Info and ReqIF Header classes that hold the
    ReqIF XML metadata and the ReqIF file identification data.
    """
    namespace_info = ReqIFNamespaceInfo.create_default()

    reqif_header = ReqIFReqIFHeader(
        identifier=create_uuid(),
        creation_time=date_now,
        repository_id="https://github.com/strictdoc-project/reqif",
        req_if_tool_id="Python ReqIF library",
        req_if_version="1.0",
        source_tool_id="Python script",
        title=(
            "An example ReqIF file created from Excel and Python objects "
            "using reqif library"
        ),
    )

    reqif_lookup = ReqIFObjectLookup.empty()

    """
    Create the final ReqIF bundle from all the objects created so far.
    """
    bundle = ReqIFBundle(
        namespace_info=namespace_info,
        req_if_header=reqif_header,
        core_content=core_content,
        tool_extensions_tag_exists=False,
        lookup=reqif_lookup,
        exceptions=[],
    )

    """
    Use ReqIF Unparser to convert the ReqIF bundle from Python objects to an XML string.
    """
    reqif_string_content: str = ReqIFUnparser.unparse(bundle)
    return reqif_string_content


def convert_reqif_to_excel(path_to_reqif: str, path_to_output_dir: str) -> None:
    assert os.path.isfile(path_to_reqif), path_to_reqif

    input_file_name = Path(path_to_reqif).stem

    wb = Workbook()
    ws = wb.active

    reqif_bundle: ReqIFBundle = ReqIFParser.parse(path_to_reqif)
    reqif_schema = ReqIFSchema(reqif_bundle)

    document_dict = []

    # For now, only one document per ReqIF is supported.
    specification: ReqIFSpecification = reqif_bundle.core_content.req_if_content.specifications[0]

    for current_hierarchy_ in reqif_bundle.iterate_specification_hierarchy(specification):
        spec_object = reqif_bundle.get_spec_object_by_ref(
            current_hierarchy_.spec_object
        )
        row_dict = _convert_spec_object_to_node(spec_object, reqif_bundle, reqif_schema, 1)
        document_dict.append(row_dict)

    for row_dict_ in document_dict:
        values_no_xhtml = []
        for value_ in row_dict_.values():
            parser = etree.XMLParser(recover=True)
            tree = etree.XML(value_, parser)

            # Extract text content, ignoring the tags.
            text_content = ''.join(tree.itertext())

            values_no_xhtml.append(text_content)

        ws.append(values_no_xhtml)

    """
    Writing the workbook to an output file.
    """
    path_to_output_dir = path_to_output_dir
    Path(path_to_output_dir).mkdir(exist_ok=True)

    path_to_output_file = os.path.join(
        path_to_output_dir, f"{input_file_name}.xlsx"
    )
    wb.save(path_to_output_file)


def main():
    main_parser = argparse.ArgumentParser()

    main_parser.add_argument(
        "input_file", type=str, help="Path to the input ReqIF file."
    )
    main_parser.add_argument(
        "--output-dir",
        type=str,
        help="Path to the output dir.",
        default="output/",
    )
    main_parser.add_argument(
        "--stdout",
        help="Makes the script write the output ReqIF to standard output.",
        action="store_true",
    )
    main_parser.add_argument(
        "--no-filesystem",
        help=(
            "Disables writing to the file system. "
            "Should be used in combination with --stdout."
        ),
        action="store_true",
    )

    args = main_parser.parse_args()
    input_file: str = args.input_file
    input_file_name = Path(input_file).stem
    should_use_file_system: bool = not args.no_filesystem
    should_use_stdout: bool = args.stdout

    if ".xlsx" in input_file:
        reqif_string_content: str = convert_excel_to_reqif(input_file)

        if should_use_file_system:
            path_to_output_dir = args.output_dir
            Path(path_to_output_dir).mkdir(exist_ok=True)

            path_to_output_file = os.path.join(
                path_to_output_dir, f"{input_file_name}.reqif"
            )
            with open(path_to_output_file, "w") as output_reqif_file:
                output_reqif_file.write(reqif_string_content)

        if should_use_stdout:
            print(reqif_string_content)  # noqa: T201
    elif ".reqif" in input_file:
        convert_reqif_to_excel(input_file, args.output_dir)


main()
